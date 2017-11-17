#! python3
# work_attendance.py 帮助助理整合考勤指纹系统和OA系统的考勤数据，生成每月的最终考勤报表

import openpyxl
from openpyxl.styles import Font,Color,Fill
from openpyxl.styles.colors import RED
from openpyxl.workbook import Workbook
from openpyxl.styles import numbers
import time
import datetime
import logging
import os
from openpyxl.chart import BarChart, Series, Reference, BarChart3D  

logging.basicConfig( filename='myProgramLog.txt', level= logging.DEBUG, format=' %(asctime) s - %(levelname) s - %(message) s')

def makechart(self, title, pos, width, height, col1, row1, col2, row2, col3, row3, row4):  
    ''''':param title:图表名 
              pos:图表位置 
              width:图表宽度 
              height:图表高度 
    '''  
    data = Reference(self.ws, min_col=col1, min_row=row1, max_col=col2, max_row=row2)  
    cat = Reference(self.ws, min_col=col3, min_row=row3, max_row=row4)  
    chart = BarChart3D()  
    chart.title = title  
    chart.width = width  
    chart.height = height  
    chart.add_data(data=data, titles_from_data=True)  
    chart.set_categories(cat)  
    self.ws.add_chart(chart, pos)  
    self.wb.save(self.filename)  

currentDir=os.getcwd()
fileList=os.listdir(currentDir)
print(fileList)
"""
fmark = 0
wmark = 0
qmark = 0
jmark = 0
cmark = 0
"""
table_cnt = 0
ws_col = 4

for fileName in fileList:
    sheetName=fileName.split('.')
    #print(peformanceSheet)
    if sheetName[1] == 'xlsx' and '研制任务书需求分析调查表' in sheetName[0]:
        if table_cnt==0:
            getNameArray=sheetName[0].split('-')
            print(getNameArray)
            tableOpen=openpyxl.load_workbook(str(currentDir+'\\'+fileName),data_only=True)
            
            curSheet_delphi=tableOpen.worksheets[0]
            print(curSheet_delphi.title)
            table_cnt+=1
            
            wb = Workbook()
            ws1 = wb.create_sheet(0)
            ws1.title = u'功能匹配度'
            ws1.cell(row=1,column=1).value='No.'
            ws1.cell(row=1,column=2).value='需求类别'
            ws1.cell(row=1,column=3).value='客户需求'
            ws1.cell(row=1,column=4).value='第'+ str(ws_col-3) + '位专家'
            ws1['A1'].style= 'Accent3'
            ws1['B1'].style= 'Accent3'
            ws1['C1'].style= 'Accent3'
            ws1['D1'].style= 'Accent2'
            ws1.column_dimensions['C'].width = 50
            ws1.column_dimensions['C'].width = 70
            ws2 = wb.create_sheet(0)
            ws2.title = u'开发难度'
            ws2.cell(row=1,column=1).value='No.'
            ws2.cell(row=1,column=2).value='需求类别'
            ws2.cell(row=1,column=3).value='客户需求'
            ws2.cell(row=1,column=4).value='第'+ str(ws_col-3) + '位专家'
            ws2['A1'].style= 'Accent3'
            ws2['B1'].style= 'Accent3'
            ws2['C1'].style= 'Accent3'
            ws2['D1'].style= 'Accent2'
            ws2.column_dimensions['C'].width = 50
            ws2.column_dimensions['C'].width = 70
            ws3 = wb.create_sheet(0)
            ws3.title = u'开发工作量'
            ws3.cell(row=1,column=1).value='No.'
            ws3.cell(row=1,column=2).value='需求类别'
            ws3.cell(row=1,column=3).value='客户需求'
            ws3.cell(row=1,column=4).value='第'+ str(ws_col-3) + '位专家'
            ws3['A1'].style= 'Accent3'
            ws3['B1'].style= 'Accent3'
            ws3['C1'].style= 'Accent3'
            ws3['D1'].style= 'Accent2'
            ws3.column_dimensions['C'].width = 50
            ws3.column_dimensions['C'].width = 70
            ws4 = wb.create_sheet(0)
            ws4.title = u'最显著差异'
            ws4.cell(row=1,column=1).value='No.'
            ws4.cell(row=1,column=2).value='需求类别'
            ws4.cell(row=1,column=3).value='客户需求'
            ws4.cell(row=1,column=4).value='第'+ str(ws_col-3) + '位专家'
            ws4['A1'].style= 'Accent3'
            ws4['B1'].style= 'Accent3'
            ws4['C1'].style= 'Accent3'
            ws4['D1'].style= 'Accent2'
            ws4.column_dimensions['C'].width = 50
            ws4.column_dimensions['C'].width = 70
            ws_line = 2
            print('第'+ str(ws_col-3) + '位专家')
            for row_line in range(2,curSheet_delphi.max_row+1):     #skip the first line
                #功能匹配度
                ws1.cell(row=ws_line,column=1).value = curSheet_delphi.cell(row= row_line, column=1).value
                ws1.cell(row=ws_line,column=2).value = curSheet_delphi.cell(row= row_line, column=2).value
                ws1.cell(row=ws_line,column=3).value = curSheet_delphi.cell(row= row_line, column=3).value
                ws1.cell(row=ws_line,column=ws_col).number_format = numbers.FORMAT_PERCENTAGE;
                if curSheet_delphi.cell(row= row_line, column=5).value == '?' or curSheet_delphi.cell(row= row_line, column=5).value == '？'or curSheet_delphi.cell(row= row_line, column=5).value == None:
                    ws1.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=5).value
                    #print(curSheet_delphi.cell(row= row_line, column=5).value,ws1.cell(row=ws_line,column=ws_col).value, 1)
                elif curSheet_delphi.cell(row= row_line, column=5).value > 1:
                    ws1.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=5).value/100
                    #print(curSheet_delphi.cell(row= row_line, column=5).value,ws1.cell(row=ws_line,column=ws_col).value, 2)
                else:
                    ws1.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=5).value
                    #print(curSheet_delphi.cell(row= row_line, column=5).value,ws1.cell(row=ws_line,column=ws_col).value, 3)
                    
                ws2.cell(row=ws_line,column=1).value = curSheet_delphi.cell(row= row_line, column=1).value
                ws2.cell(row=ws_line,column=2).value = curSheet_delphi.cell(row= row_line, column=2).value
                ws2.cell(row=ws_line,column=3).value = curSheet_delphi.cell(row= row_line, column=3).value
                ws2.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=7).value
                    
                ws3.cell(row=ws_line,column=1).value = curSheet_delphi.cell(row= row_line, column=1).value
                ws3.cell(row=ws_line,column=2).value = curSheet_delphi.cell(row= row_line, column=2).value
                ws3.cell(row=ws_line,column=3).value = curSheet_delphi.cell(row= row_line, column=3).value
                ws3.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=8).value
                    
                ws4.cell(row=ws_line,column=1).value = curSheet_delphi.cell(row= row_line, column=1).value
                ws4.cell(row=ws_line,column=2).value = curSheet_delphi.cell(row= row_line, column=2).value
                ws4.cell(row=ws_line,column=3).value = curSheet_delphi.cell(row= row_line, column=3).value
                ws4.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=6).value
                    
                if curSheet_delphi.cell(row= row_line, column=2).value == curSheet_delphi.cell(row= row_line+1, column=2).value:
                    ws_line+=1
                   # print(str(currentDir+'\\delphi\\'+tableName))
                else:
                    tableName=curSheet_delphi.cell(row= row_line, column=2).value
                    #print(str(currentDir+'\\delphi\\'+tableName))
                    #wb.save(filename=str(currentDir+'\\delphi\\'+tableName)) 
                    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
                    #wb.remove_sheet(    )
                    wb.save(filename=str(currentDir+'\\'+tableName+'.xlsx')) 
                    wb = Workbook()
                    ws1 = wb.create_sheet(0)
                    ws1.title = u'功能匹配度'
                    ws1.cell(row=1,column=1).value='No.'
                    ws1.cell(row=1,column=2).value='需求类别'
                    ws1.cell(row=1,column=3).value='客户需求'
                    ws1.cell(row=1,column=4).value='第'+ str(ws_col-3) + '位专家'
                    ws1['A1'].style= 'Accent3'
                    ws1['B1'].style= 'Accent3'
                    ws1['C1'].style= 'Accent3'
                    ws1['D1'].style= 'Accent2'
                    ws1.column_dimensions['C'].width = 50
                    ws1.column_dimensions['C'].width = 70
                    ws2 = wb.create_sheet(0)
                    ws2.title = u'开发难度'
                    ws2.cell(row=1,column=1).value='No.'
                    ws2.cell(row=1,column=2).value='需求类别'
                    ws2.cell(row=1,column=3).value='客户需求'
                    ws2.cell(row=1,column=4).value='第'+ str(ws_col-3) + '位专家'
                    ws2['A1'].style= 'Accent3'
                    ws2['B1'].style= 'Accent3'
                    ws2['C1'].style= 'Accent3'
                    ws2['D1'].style= 'Accent2'
                    ws2.column_dimensions['C'].width = 50
                    ws2.column_dimensions['C'].width = 70
                    ws3 = wb.create_sheet(0)
                    ws3.title = u'开发工作量'
                    ws3.cell(row=1,column=1).value='No.'
                    ws3.cell(row=1,column=2).value='需求类别'
                    ws3.cell(row=1,column=3).value='客户需求'
                    ws3.cell(row=1,column=4).value='第'+ str(ws_col-3) + '位专家'
                    ws3['A1'].style= 'Accent3'
                    ws3['B1'].style= 'Accent3'
                    ws3['C1'].style= 'Accent3'
                    ws3['D1'].style= 'Accent2'
                    ws3.column_dimensions['C'].width = 50
                    ws3.column_dimensions['C'].width = 70
                    ws4 = wb.create_sheet(0)
                    ws4.title = u'最显著差异'
                    ws4.cell(row=1,column=1).value='No.'
                    ws4.cell(row=1,column=2).value='需求类别'
                    ws4.cell(row=1,column=3).value='客户需求'
                    ws4.cell(row=1,column=4).value='第'+ str(ws_col-3) + '位专家'
                    ws4['A1'].style= 'Accent3'
                    ws4['B1'].style= 'Accent3'
                    ws4['C1'].style= 'Accent3'
                    ws4['D1'].style= 'Accent2'
                    ws4.column_dimensions['C'].width = 50
                    ws4.column_dimensions['C'].width = 70
                    ws_line = 2
                    #ws_col = 3

        else:
            getNameArray=sheetName[0].split('-')
            print(getNameArray)
            tableOpen=openpyxl.load_workbook(str(currentDir+'\\'+fileName),data_only=True)
            
            curSheet_delphi=tableOpen.worksheets[0]
            print(curSheet_delphi.title)
            table_cnt+=1
            ws_line = 2
            ws_col+=1
            wb = openpyxl.load_workbook(str(currentDir+'\\'+str(curSheet_delphi.cell(row= 2, column=2).value)+'.xlsx'),data_only=True)
            ws1 = wb.get_sheet_by_name('功能匹配度')
            ws1.cell(row=1,column=ws_col).value='第'+ str(ws_col-3) + '位专家'
            ws1.cell(row=1,column=ws_col).style='Accent2'
            ws2 = wb.get_sheet_by_name('开发难度')
            ws2.cell(row=1,column=ws_col).value='第'+ str(ws_col-3) + '位专家'
            ws2.cell(row=1,column=ws_col).style='Accent2'
            ws3 = wb.get_sheet_by_name('开发工作量')
            ws3.cell(row=1,column=ws_col).value='第'+ str(ws_col-3) + '位专家'
            ws3.cell(row=1,column=ws_col).style='Accent2'
            ws4 = wb.get_sheet_by_name('最显著差异')
            ws4.cell(row=1,column=ws_col).value='第'+ str(ws_col-3) + '位专家'
            ws4.cell(row=1,column=ws_col).style='Accent2'
            print('第'+ str(ws_col-3) + '位专家')
            
            for row_line in range(2,curSheet_delphi.max_row+1):     #skip the first line
                #功能匹配度
                #ws1.cell(row=ws_line,column=1).value = curSheet_delphi.cell(row= row_line, column=1).value
                #ws1.cell(row=ws_line,column=2).value = curSheet_delphi.cell(row= row_line, column=2).value
                ws1.cell(row=ws_line,column=ws_col).number_format = numbers.FORMAT_PERCENTAGE;
                
                if curSheet_delphi.cell(row= row_line, column=5).value == '?' or curSheet_delphi.cell(row= row_line, column=5).value == '？'or curSheet_delphi.cell(row= row_line, column=5).value == None:
                    ws1.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=5).value
                    #print(curSheet_delphi.cell(row= row_line, column=5).value,ws1.cell(row=ws_line,column=ws_col).value, 1)
                elif curSheet_delphi.cell(row= row_line, column=5).value > 1:
                    ws1.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=5).value/100
                    #print(curSheet_delphi.cell(row= row_line, column=5).value,ws1.cell(row=ws_line,column=ws_col).value,2)
                else:
                    ws1.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=5).value
                    #print(curSheet_delphi.cell(row= row_line, column=5).value,ws1.cell(row=ws_line,column=ws_col).value,3)
                #ws1.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=5).value
                    
                #ws2.cell(row=ws_line,column=1).value = curSheet_delphi.cell(row= row_line, column=1).value
                #ws2.cell(row=ws_line,column=2).value = curSheet_delphi.cell(row= row_line, column=2).value
                ws2.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=7).value
                    
                #ws3.cell(row=ws_line,column=1).value = curSheet_delphi.cell(row= row_line, column=1).value
                #ws3.cell(row=ws_line,column=2).value = curSheet_delphi.cell(row= row_line, column=2).value
                ws3.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=8).value
                    
                #ws4.cell(row=ws_line,column=1).value = curSheet_delphi.cell(row= row_line, column=1).value
                #ws4.cell(row=ws_line,column=2).value = curSheet_delphi.cell(row= row_line, column=2).value
                ws4.cell(row=ws_line,column=ws_col).value = curSheet_delphi.cell(row= row_line, column=6).value
                    
                if curSheet_delphi.cell(row= row_line, column=2).value == curSheet_delphi.cell(row= row_line+1, column=2).value:
                    ws_line+=1
                   # print(str(currentDir+'\\delphi\\'+tableName))
                else:
                    tableName=curSheet_delphi.cell(row= row_line, column=2).value
                    #print(str(currentDir+'\\delphi\\'+tableName))
                    #wb.save(filename=str(currentDir+'\\delphi\\'+tableName)) 
                    wb.save(filename=str(currentDir+'\\'+tableName+'.xlsx')) 
                    
                    if row_line < curSheet_delphi.max_row and curSheet_delphi.cell(row= row_line+1, column=2).value != None : 
                        wb = openpyxl.load_workbook(str(currentDir+'\\'+str(curSheet_delphi.cell(row= row_line+1, column=2).value)+'.xlsx'),data_only=True)
                        ws1 = wb.get_sheet_by_name('功能匹配度')
                        ws1.cell(row=1,column=ws_col).value='第'+ str(ws_col-3) + '位专家'
                        ws1.cell(row=1,column=ws_col).style='Accent2'
                        ws2 = wb.get_sheet_by_name('开发难度')
                        ws2.cell(row=1,column=ws_col).value='第'+ str(ws_col-3) + '位专家'
                        ws2.cell(row=1,column=ws_col).style='Accent2'
                        ws3 = wb.get_sheet_by_name('开发工作量')
                        ws3.cell(row=1,column=ws_col).value='第'+ str(ws_col-3) + '位专家'
                        ws3.cell(row=1,column=ws_col).style='Accent2'
                        ws4 = wb.get_sheet_by_name('最显著差异')
                        ws4.cell(row=1,column=ws_col).value='第'+ str(ws_col-3) + '位专家'
                        ws4.cell(row=1,column=ws_col).style='Accent2'
                        ws_line = 2
                    
            print(row_line, curSheet_delphi.max_row)        
            #else end
        
        
        
        
print(table_cnt)        
"""            


ft = Font(bold=True)

fingerSheet['G1']= '请假开始时间'
fingerSheet['G1'].font=ft

fingerSheet['H1']= '请假结束时间'
fingerSheet['H1'].font=ft

fingerSheet['I1']= '请假类型'
fingerSheet['I1'].font=ft

fingerSheet['J1']= '请假原因'
fingerSheet['J1'].font=ft

fingerSheet['K1']= '出差开始时间'
fingerSheet['K1'].font=ft

fingerSheet['L1']= '出差结束时间'
fingerSheet['L1'].font=ft

fingerSheet['M1']= '外勤开始时间'
fingerSheet['M1'].font=ft

fingerSheet['N1']= '外勤结束时间'
fingerSheet['N1'].font=ft

fingerSheet['O1']= '加班开始时间'
fingerSheet['O1'].font=ft

fingerSheet['P1']= '加班结束时间'
fingerSheet['P1'].font=ft

fingerSheet['Q1']= '补卡开始时间'
fingerSheet['Q1'].font=ft

fingerSheet['R1']= '补卡结束时间'
fingerSheet['R1'].font=ft

fingerSheet.freeze_panes = 'F2'


"""